import numpy as np
import pandas as pd
import pytest
import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from shared import (
    filter_targets,
    group_ct_data,
    welch_ttest,
    one_sample_ttest,
    correct_pvalues,
    p_to_star,
    export_to_csv,
    export_to_excel,
    _split_sample_name,
    _transform_quantstudio_df,
    _is_quantstudio_export,
    import_file,
    one_way_anova,
    kruskal_wallis,
    tukey_hsd_posthoc,
    pairwise_welch_posthoc,
    dunns_posthoc,
    two_way_anova,
    _lookup_posthoc_pval,
)


@pytest.fixture
def sample_df():
    return pd.DataFrame({
        "Group": ["Grp1", "Grp1", "Grp1", "Grp1"],
        "Condition": ["A", "A", "B", "B"],
        "Target Name": ["Gene1", "Gene1", "Gene1", "Gene1"],
        "CT": [20.0, 21.0, 22.0, 23.0],
    })


class TestFilterTargets:
    def test_exact_match(self, sample_df):
        result = filter_targets(sample_df, "Gene1")
        assert len(result) == 4

    def test_no_match(self, sample_df):
        result = filter_targets(sample_df, "GeneX")
        assert len(result) == 0


class TestGroupCtData:
    def test_groups_correctly(self, sample_df):
        result = group_ct_data(sample_df)
        assert "Gene1_Grp1_A" in result
        assert "Gene1_Grp1_B" in result
        assert len(result["Gene1_Grp1_A"]) == 2
        assert len(result["Gene1_Grp1_B"]) == 2


class TestWelchTtest:
    def test_identical_groups(self):
        g1 = np.array([20.0, 20.0, 20.0])
        g2 = np.array([20.0, 20.0, 20.0])
        t, p = welch_ttest(g1, g2)
        assert np.isnan(t)  # 0/0 with no variance

    def test_different_groups(self):
        g1 = np.array([10.0, 11.0, 12.0])
        g2 = np.array([20.0, 21.0, 22.0])
        t, p = welch_ttest(g1, g2)
        assert p < 0.05

    def test_single_replicate_returns_nan(self):
        t, p = welch_ttest(np.array([10.0]), np.array([20.0]))
        assert np.isnan(p)


class TestOneSampleTtest:
    def test_mean_zero(self):
        vals = np.array([0.0, 0.0, 0.0])
        t, p = one_sample_ttest(vals, popmean=0)
        assert np.isnan(t)  # 0/0

    def test_mean_nonzero(self):
        vals = np.array([5.0, 5.1, 4.9])
        t, p = one_sample_ttest(vals, popmean=0)
        assert p < 0.05


class TestCorrectPvalues:
    def test_none_correction(self):
        p = np.array([0.01, 0.04, 0.1])
        result = correct_pvalues(p, method="none")
        np.testing.assert_array_equal(result, p)

    def test_bonferroni(self):
        p = np.array([0.01, 0.04, 0.1])
        result = correct_pvalues(p, method="bonferroni")
        np.testing.assert_allclose(result, [0.03, 0.12, 0.3])

    def test_bonferroni_caps_at_one(self):
        p = np.array([0.5, 0.8])
        result = correct_pvalues(p, method="bonferroni")
        assert all(r <= 1.0 for r in result)

    def test_bh_correction(self):
        p = np.array([0.01, 0.04, 0.1])
        result = correct_pvalues(p, method="bh")
        # BH should be less aggressive than Bonferroni
        bonf = correct_pvalues(p, method="bonferroni")
        assert all(result <= bonf + 1e-10)


class TestPToStar:
    def test_significant(self):
        assert p_to_star(0.0001) == "***"
        assert p_to_star(0.005) == "**"
        assert p_to_star(0.03) == "*"
        assert p_to_star(0.1) == "ns"

    def test_nan(self):
        assert p_to_star(np.nan) == ""


class TestExport:
    def test_csv_export(self, sample_df):
        result = export_to_csv(sample_df)
        assert isinstance(result, bytes)
        assert b"Target Name" in result

    def test_excel_export(self, sample_df):
        result = export_to_excel(sample_df)
        assert isinstance(result, bytes)
        assert len(result) > 0


class TestSplitSampleName:
    def test_simple_split(self):
        group, condition = _split_sample_name("WT_treated")
        assert group == "WT"
        assert condition == "treated"

    def test_underscore_in_group(self):
        group, condition = _split_sample_name("anti_Flag_enrich")
        assert group == "anti_Flag"
        assert condition == "enrich"

    def test_no_underscore_raises(self):
        with pytest.raises(ValueError, match="no underscore"):
            _split_sample_name("nosplit")

    def test_single_underscore(self):
        group, condition = _split_sample_name("A_B")
        assert group == "A"
        assert condition == "B"


class TestTransformQuantstudioDf:
    @pytest.fixture
    def raw_df(self):
        return pd.DataFrame({
            "Sample Name": ["WT_ctrl", "WT_ctrl", "KO_treated"],
            "Target Name": ["Gene1", "Gene2", "Gene1"],
            "CT": [20.0, 21.0, 22.0],
            "Task": ["Unknown", "Unknown", "Unknown"],
        })

    def test_columns_present(self, raw_df):
        result = _transform_quantstudio_df(raw_df)
        assert set(["Group", "Condition", "Target Name", "CT", "Biological Replicate"]).issubset(result.columns)

    def test_sample_name_removed(self, raw_df):
        result = _transform_quantstudio_df(raw_df)
        assert "Sample Name" not in result.columns

    def test_group_condition_values(self, raw_df):
        result = _transform_quantstudio_df(raw_df)
        assert list(result["Group"]) == ["WT", "WT", "KO"]
        assert list(result["Condition"]) == ["ctrl", "ctrl", "treated"]

    def test_task_renamed_to_biological_replicate(self, raw_df):
        result = _transform_quantstudio_df(raw_df)
        assert "Biological Replicate" in result.columns
        assert "Task" not in result.columns

    def test_input_not_mutated(self, raw_df):
        original_cols = list(raw_df.columns)
        _transform_quantstudio_df(raw_df)
        assert list(raw_df.columns) == original_cols


class TestIsQuantstudioExport:
    def test_csv_returns_false(self, tmp_path):
        f = tmp_path / "data.csv"
        f.write_text("a,b\n1,2\n")
        assert _is_quantstudio_export(str(f)) is False

    def test_xlsx_without_results_sheet(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        path = str(tmp_path / "no_results.xlsx")
        wb.save(path)
        assert _is_quantstudio_export(path) is False

    def test_xlsx_with_results_sheet(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Results"
        path = str(tmp_path / "with_results.xlsx")
        wb.save(path)
        assert _is_quantstudio_export(path) is True


class TestOneWayAnova:
    def test_significant(self):
        g1 = np.array([10.0, 11.0, 10.5])
        g2 = np.array([20.0, 21.0, 20.5])
        g3 = np.array([30.0, 31.0, 30.5])
        F, p = one_way_anova(g1, g2, g3)
        assert not np.isnan(F)
        assert p < 0.05

    def test_not_significant(self):
        g1 = np.array([10.0, 10.1, 10.2])
        g2 = np.array([10.1, 10.2, 10.3])
        g3 = np.array([10.2, 10.3, 10.4])
        F, p = one_way_anova(g1, g2, g3)
        assert not np.isnan(F)
        assert p > 0.05

    def test_insufficient_data_returns_nan(self):
        g1 = np.array([10.0])  # only 1 value — too few for valid group
        F, p = one_way_anova(g1)
        assert np.isnan(F)
        assert np.isnan(p)

    def test_single_valid_group_returns_nan(self):
        g1 = np.array([10.0, 11.0])
        F, p = one_way_anova(g1)
        assert np.isnan(F)
        assert np.isnan(p)


class TestKruskalWallis:
    def test_significant(self):
        g1 = np.array([1.0, 2.0, 1.5])
        g2 = np.array([10.0, 11.0, 10.5])
        H, p = kruskal_wallis(g1, g2)
        assert not np.isnan(H)
        assert p < 0.05

    def test_not_significant(self):
        g1 = np.array([10.0, 10.1, 10.2])
        g2 = np.array([10.1, 10.2, 10.3])
        H, p = kruskal_wallis(g1, g2)
        assert not np.isnan(H)
        assert p > 0.05

    def test_insufficient_groups_returns_nan(self):
        g1 = np.array([10.0, 11.0])
        H, p = kruskal_wallis(g1)
        assert np.isnan(H)
        assert np.isnan(p)


class TestTukeyHsdPosthoc:
    def test_correct_columns(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
            "C": np.array([30.0, 30.5, 30.2]),
        }
        result = tukey_hsd_posthoc(groups)
        assert set(result.columns) == {"group1", "group2", "p_adj", "significance"}

    def test_number_of_pairs(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
            "C": np.array([30.0, 30.5, 30.2]),
        }
        result = tukey_hsd_posthoc(groups)
        # 3 groups → 3 pairs
        assert len(result) == 3

    def test_significant_pair(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([30.0, 30.5, 30.2]),
        }
        result = tukey_hsd_posthoc(groups)
        assert result.iloc[0]["p_adj"] < 0.05

    def test_insufficient_data_returns_empty(self):
        groups = {"A": np.array([10.0])}
        result = tukey_hsd_posthoc(groups)
        assert result.empty


class TestPairwiseWelchPosthoc:
    def test_correct_columns(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
        }
        result = pairwise_welch_posthoc(groups)
        assert set(result.columns) == {"group1", "group2", "p_adj", "significance"}

    def test_bonferroni_applied(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
            "C": np.array([30.0, 30.5, 30.2]),
        }
        result = pairwise_welch_posthoc(groups)
        # 3 pairs — Bonferroni multiplies by 3; at least one pair should be significant
        assert len(result) == 3
        assert all(result["p_adj"] <= 1.0)

    def test_significant_pair(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([30.0, 30.5, 30.2]),
        }
        result = pairwise_welch_posthoc(groups)
        assert result.iloc[0]["p_adj"] < 0.05


class TestDunnsPosthoc:
    def test_correct_columns(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
        }
        result = dunns_posthoc(groups)
        assert set(result.columns) == {"group1", "group2", "p_adj", "significance"}

    def test_bonferroni_correction(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
            "C": np.array([30.0, 30.5, 30.2]),
        }
        result = dunns_posthoc(groups, correction="bonferroni")
        assert len(result) == 3
        assert all(result["p_adj"] <= 1.0)

    def test_bh_correction(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
            "C": np.array([30.0, 30.5, 30.2]),
        }
        result_bh = dunns_posthoc(groups, correction="bh")
        result_bonf = dunns_posthoc(groups, correction="bonferroni")
        # BH should be <= Bonferroni (less conservative)
        assert all(result_bh["p_adj"].values <= result_bonf["p_adj"].values + 1e-10)

    def test_dunn_bonf_alias(self):
        groups = {
            "A": np.array([10.0, 10.5, 10.2]),
            "B": np.array([20.0, 20.5, 20.2]),
        }
        result = dunns_posthoc(groups, correction="dunn_bonf")
        assert len(result) == 1

    def test_significant_pair(self):
        # Need n>=5 per group: Mann-Whitney min p-value with n=3 is 0.1 (combinatorics limit)
        groups = {
            "A": np.array([1.0, 1.1, 1.2, 1.3, 1.4]),
            "B": np.array([100.0, 100.1, 100.2, 100.3, 100.4]),
        }
        result = dunns_posthoc(groups)
        assert result.iloc[0]["p_adj"] < 0.05


class TestTwoWayAnova:
    @pytest.fixture
    def factorial_df(self):
        # 2x2 factorial: Group (WT/KO) × Condition (ctrl/treated)
        data = []
        np.random.seed(42)
        for grp, base in [("WT", 20.0), ("KO", 25.0)]:
            for cond, offset in [("ctrl", 0.0), ("treated", 5.0)]:
                for _ in range(4):
                    data.append({
                        "Group": grp,
                        "Condition": cond,
                        "CT": base + offset + np.random.normal(0, 0.5),
                    })
        return pd.DataFrame(data)

    def test_correct_columns(self, factorial_df):
        result = two_way_anova(factorial_df, "CT", "Group", "Condition")
        assert set(result.columns) == {"Factor", "F", "p_value", "significance"}

    def test_factor_names_in_result(self, factorial_df):
        result = two_way_anova(factorial_df, "CT", "Group", "Condition")
        factors = result["Factor"].tolist()
        assert "Group" in factors
        assert "Condition" in factors
        assert "Group:Condition" in factors

    def test_significant_factors(self, factorial_df):
        result = two_way_anova(factorial_df, "CT", "Group", "Condition")
        group_row = result[result["Factor"] == "Group"].iloc[0]
        cond_row = result[result["Factor"] == "Condition"].iloc[0]
        assert group_row["p_value"] < 0.05
        assert cond_row["p_value"] < 0.05

    def test_insufficient_data_returns_empty(self):
        df = pd.DataFrame({
            "Group": ["WT"],
            "Condition": ["ctrl"],
            "CT": [20.0],
        })
        result = two_way_anova(df, "CT", "Group", "Condition")
        assert result.empty


class TestLookupPosthocPval:
    @pytest.fixture
    def ph_df(self):
        return pd.DataFrame({
            "group1": ["A", "A", "B"],
            "group2": ["B", "C", "C"],
            "p_adj": [0.01, 0.5, 0.03],
            "significance": ["*", "ns", "*"],
        })

    def test_forward_order(self, ph_df):
        p = _lookup_posthoc_pval(ph_df, "A", "B")
        assert abs(p - 0.01) < 1e-10

    def test_reverse_order(self, ph_df):
        p = _lookup_posthoc_pval(ph_df, "B", "A")
        assert abs(p - 0.01) < 1e-10

    def test_missing_pair_returns_nan(self, ph_df):
        p = _lookup_posthoc_pval(ph_df, "A", "D")
        assert np.isnan(p)

    def test_empty_df_returns_nan(self):
        ph_df = pd.DataFrame(columns=["group1", "group2", "p_adj", "significance"])
        p = _lookup_posthoc_pval(ph_df, "A", "B")
        assert np.isnan(p)

    def test_none_returns_nan(self):
        p = _lookup_posthoc_pval(None, "A", "B")
        assert np.isnan(p)


def _build_quantstudio_xlsx(tmp_path, rows):
    """Write a minimal QuantStudio xlsx: 46 blank header rows then data rows."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    # 46 blank header rows
    for _ in range(46):
        ws.append([])
    # Header row (row 47 = skiprows=46 reads from here)
    ws.append(["Sample Name", "Target Name", "CT", "Task"])
    for row in rows:
        ws.append(row)
    path = str(tmp_path / "instrument.xlsx")
    wb.save(path)
    return path


class TestImportFileQuantstudioFormat:
    def test_expected_columns(self, tmp_path):
        path = _build_quantstudio_xlsx(tmp_path, [
            ["WT_ctrl", "Gene1", 20.0, "Unknown"],
        ])
        df = import_file(path)
        assert set(["Group", "Condition", "Target Name", "CT", "Biological Replicate"]).issubset(df.columns)

    def test_group_condition_parsed(self, tmp_path):
        path = _build_quantstudio_xlsx(tmp_path, [
            ["anti_Flag_enrich", "Gene1", 20.0, "Unknown"],
        ])
        df = import_file(path)
        assert df["Group"].iloc[0] == "anti_Flag"
        assert df["Condition"].iloc[0] == "enrich"

    def test_undetermined_row_dropped(self, tmp_path):
        path = _build_quantstudio_xlsx(tmp_path, [
            ["WT_ctrl", "Gene1", 20.0, "Unknown"],
            ["WT_ctrl", "Gene1", "Undetermined", "Unknown"],
        ])
        df = import_file(path)
        assert len(df) == 1

    def test_biological_replicate_populated(self, tmp_path):
        path = _build_quantstudio_xlsx(tmp_path, [
            ["WT_ctrl", "Gene1", 20.0, "Unknown"],
        ])
        df = import_file(path)
        assert df["Biological Replicate"].iloc[0] == "Unknown"
